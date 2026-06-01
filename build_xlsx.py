# -*- coding: utf-8 -*-
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = list(csv.DictReader(open("padron_aeronaves.csv", encoding="utf-8")))
# Tenencia derivada: Charter si el organismo lo indica; si no, Propia
for r in rows:
    r["tenencia"] = "Charter" if "charter" in (r.get("organismo","").lower()) else "Propia"

cols  = ["provincia","organismo","matricula","hex","tipo","categoria","estado","tenencia","confianza","titular_registrado","fuente","notas"]
heads = ["Provincia","Organismo / Titular","Matrícula","Hex ICAO","Tipo de aeronave",
         "Categoría","Estado","Tenencia","Confianza","Titular registrado (BD)","Fuente","Notas"]
widths = [16,32,12,11,24,12,16,11,13,28,16,38]
NCOL = len(cols)
LAST = get_column_letter(NCOL)

AZUL = "1F4E79"; GRIS = "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
title_f = Font(name="Arial", size=14, bold=True, color="FFFFFF")
head_f  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
cell_f  = Font(name="Arial", size=10)
mono_f  = Font(name="Arial", size=10, bold=True)

conf_fill = {"VERIF. ANAC":"9BC2E6","CRUCE BD":"B4C6E7","ALTA":"C6EFCE","MEDIA":"FFEB9C","BAJA":"F8CBAD"}
conf_font = {"VERIF. ANAC":"1F4E79","CRUCE BD":"1F4E79","ALTA":"006100","MEDIA":"9C6500","BAJA":"9C3C00"}
ten_fill  = {"Propia":"C6EFCE","Charter":"FCE4D6"}
ten_font  = {"Propia":"006100","Charter":"C55A11"}

def render(ws, data, title, subtitle, heads_arg=None):
    hh = heads_arg or heads
    ws.merge_cells(f"A1:{LAST}1")
    ws["A1"] = title
    ws["A1"].font = title_f
    ws["A1"].fill = PatternFill("solid", fgColor=AZUL)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{LAST}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18
    hr = 3
    for j,h in enumerate(hh,1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font = head_f; c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[hr].height = 30
    r = hr+1
    for row in data:
        for j,key in enumerate(cols,1):
            c = ws.cell(row=r, column=j, value=row.get(key,""))
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=(key in ("organismo","notas")), indent=1)
            c.font = mono_f if key in ("matricula","hex") else cell_f
            if key=="confianza" and row.get(key) in conf_fill:
                c.fill = PatternFill("solid", fgColor=conf_fill[row[key]])
                c.font = Font(name="Arial", size=10, bold=True, color=conf_font[row[key]])
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif key=="tenencia" and row.get(key) in ten_fill:
                c.fill = PatternFill("solid", fgColor=ten_fill[row[key]])
                c.font = Font(name="Arial", size=10, bold=True, color=ten_font[row[key]])
                c.alignment = Alignment(horizontal="center", vertical="center")
        if r % 2 == 0:
            for j in range(1,NCOL+1):
                k = cols[j-1]
                if not (k in ("confianza","tenencia") and row.get(k) in conf_fill | ten_fill):
                    ws.cell(row=r,column=j).fill = PatternFill("solid", fgColor=GRIS)
        r += 1
    for j,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hr}:{LAST}{r-1}"

wb = Workbook()
propias  = [r for r in rows if r["tenencia"]=="Propia"]
charters = [r for r in rows if r["tenencia"]=="Charter"]

ws = wb.active; ws.title = "Padrón completo"
render(ws, rows, "Padrón de aeronaves de gobiernos provinciales — Argentina",
       f"{len(rows)} aeronaves · {len(propias)} flota propia · {len(charters)} charter/contratado. Tenencia: Propia = matrícula del Estado provincial; Charter = aeronave de una empresa usada por la gobernación.")

render(wb.create_sheet("Flota propia"), propias, "Flota propia provincial",
       "Aeronaves cuya matrícula pertenece (o se atribuye) al Estado provincial. Cruzar 'Titular registrado (BD)' para detectar casos a nombre de terceros.")

render(wb.create_sheet("Charter o contratado"), charters, "Charters usados por las gobernaciones",
       "Aeronaves de empresas privadas usadas para trasladar gobernadores/funcionarios. NO son propiedad provincial; el gasto suele ser más difícil de rastrear.")

# Hoja Charters de políticos / dirigentes (registro separado, NO flota provincial)
try:
    drows = list(csv.DictReader(open("charters_dirigentes.csv", encoding="utf-8")))
except Exception:
    drows = []
for r in drows:
    r["tenencia"] = "Charter"
if drows:
    dheads = ["Ámbito / sector","Usuario (político/dirigente)","Matrícula","Hex ICAO","Tipo de aeronave",
              "Categoría","Estado","Tenencia","Confianza","Titular registrado (BD)","Fuente","Notas"]
    render(wb.create_sheet("Charter políticos-dirigentes"), drows,
           "Charters usados por políticos y dirigentes",
           "Registro SEPARADO de la flota provincial: aeronaves privadas usadas por dirigentes (AFA, gremios, etc.). Titular real en la columna 'Titular registrado (BD)'.",
           heads_arg=dheads)

# Hoja Guía
g = wb.create_sheet("Guía")
g.column_dimensions["A"].width = 22
g.column_dimensions["B"].width = 92
g.merge_cells("A1:B1")
g["A1"] = "Cómo leer y usar este padrón"
g["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
g["A1"].fill = PatternFill("solid", fgColor=AZUL)
g["A1"].alignment = Alignment(vertical="center", indent=1)
g.row_dimensions[1].height = 26
guia = [
 ("Objetivo","Registrar dónde se mueven las aeronaves de los gobiernos provinciales, con fines periodísticos y de investigación."),
 ("Tenencia","Propia = la matrícula pertenece (o se atribuye) al Estado provincial. Charter = aeronave de una empresa privada usada para trasladar al gobernador/funcionarios."),
 ("Matrícula","Prefijo LQ- = aeronave pública/estatal; LV- = privada. Muchas provincias mantienen sus aviones con matrícula LV- aunque sean del Estado."),
 ("Hex ICAO","Código de 24 bits que emite el transponder. Es la clave para rastrear la aeronave por ADS-B (airplanes.live, OpenSky, adsb.fi)."),
 ("Titular registrado (BD)","Propietario según la base de aviación (hexdb). Si NO coincide con la provincia, es una señal a investigar (charter, banco, empresa, leasing)."),
 ("Confianza","VERIF. ANAC = consulta oficial; CRUCE BD = titular confirmado en base de aviación; ALTA = matrícula confirmada en fuente; MEDIA/BAJA = a verificar."),
 ("Verificación","Registro Nacional de Aeronaves (ANAC): cad.anac.gob.ar/ConsultaPSA/ConsultasPublicas/AfectacionAeronaves"),
 ("Recolección","poller.py consulta airplanes.live por estos hex y guarda cada vuelo en vuelos.db (SQLite)."),
 ("Importante","Punto de partida, NO un padrón oficial cerrado. Toda publicación debería cruzarse con el registro ANAC y la agenda oficial del funcionario."),
]
gr = 3
for k,v in guia:
    g.cell(row=gr,column=1,value=k).font = Font(name="Arial", size=10, bold=True, color=AZUL)
    g.cell(row=gr,column=1).alignment = Alignment(vertical="top", wrap_text=True)
    g.cell(row=gr,column=2,value=v).font = Font(name="Arial", size=10)
    g.cell(row=gr,column=2).alignment = Alignment(vertical="top", wrap_text=True)
    g.row_dimensions[gr].height = 30
    gr += 1

wb.save("Padron_aeronaves_provinciales.xlsx")
print(f"OK -> {len(rows)} aeronaves | propias {len(propias)} | charters {len(charters)}")
